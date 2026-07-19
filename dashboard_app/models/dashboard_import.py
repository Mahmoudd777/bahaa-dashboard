import base64
import csv
import io
import json

from odoo import api, fields, models
from odoo.exceptions import UserError

from .dashboard_import_registry import get_target, list_targets


class DashboardImport(models.Model):
    _name = "dashboard.import"
    _description = "Dashboard Data Import"
    _order = "create_date desc"

    name = fields.Char(default="استيراد بيانات")
    user_id = fields.Many2one(
        "res.users", default=lambda self: self.env.user, required=True
    )
    date = fields.Datetime(default=fields.Datetime.now)
    attachment_ids = fields.Many2many("ir.attachment", string="Files")
    state = fields.Selection(
        selection=[
            ("uploaded", "Uploaded"),
            ("processed", "Processed"),
            ("failed", "Failed"),
        ],
        default="uploaded",
    )
    note = fields.Text()
    target_key = fields.Char(string="Import Target Key")
    target_model = fields.Char(string="Target Model")
    target_label = fields.Char(string="Target Label")
    created_count = fields.Integer(default=0)
    updated_count = fields.Integer(default=0)
    error_count = fields.Integer(default=0)
    error_log = fields.Text(string="Error Log")

    @api.model
    def get_import_targets(self, mode="normal"):
        """RPC entry: whitelisted models the current user may import."""
        if not self.env.user.sudo().dashboard_access:
            return []
        if mode == "advanced":
            if self.env.user.share or not self.env.user.has_group("base.group_user"):
                return []
            return list_targets(self.env, require_create=True)
        self.check_access_rights("create")
        return list_targets(self.sudo().env, require_create=False)

    @api.model
    def build_template_xlsx(self, target_key, template_mode="normal"):
        """Generate an Excel template for the given import target."""
        if not self.env.user.sudo().dashboard_access:
            raise UserError("Dashboard access is required.")
        import xlsxwriter  # noqa: PLC0415

        self.check_access_rights("create")
        spec = self._resolve_target(target_key, require_create=False)
        advanced = template_mode == "advanced"
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {"in_memory": True})
        sheet = workbook.add_worksheet(spec["label"][:31])
        sheet.right_to_left()

        header_fmt = workbook.add_format(
            {"bold": True, "bg_color": "#5C4B43", "font_color": "#FFFFFF", "border": 1}
        )
        cell_fmt = workbook.add_format({"border": 1})
        columns = spec["columns"]

        for col, col_spec in enumerate(columns):
            sheet.write(0, col, self._template_header(col_spec, advanced=advanced), header_fmt)
            if advanced:
                sheet.write_comment(0, col, self._template_comment(col_spec))
            sheet.set_column(col, col, 22)

        for r, sample in enumerate(spec.get("samples") or [], start=1):
            for c, value in enumerate(sample):
                if c < len(columns):
                    sheet.write(r, c, value, cell_fmt)

        workbook.close()
        output.seek(0)
        return output.read(), spec

    @api.model
    def _template_header(self, col_spec, advanced=False):
        if not advanced:
            return col_spec["header"]
        if col_spec.get("relation") and col_spec.get("relation_lookup"):
            return "%s/%s" % (col_spec["field"], col_spec["relation_lookup"])
        return col_spec["field"]

    @api.model
    def _template_comment(self, col_spec):
        notes = [col_spec["header"]]
        if col_spec.get("required"):
            notes.append("مطلوب")
        if col_spec.get("relation"):
            lookup = col_spec.get("relation_lookup") or "name"
            notes.append("سجل مرتبط: %s عبر %s" % (col_spec["relation"], lookup))
        return " | ".join(notes)

    @api.model
    def process_upload(self, target_key, file_content, filename):
        """Validate and import rows from an uploaded Excel or CSV file."""
        if not self.env.user.sudo().dashboard_access:
            raise UserError("Dashboard access is required.")
        self.check_access_rights("create")
        spec = self._resolve_target(target_key, require_create=False)
        rows = self._read_rows(file_content, filename, spec)
        if not rows:
            raise UserError("الملف لا يحتوي على بيانات للاستيراد.")

        errors = []
        prepared = []
        model = self.env[spec["model"]].sudo()

        for row_num, raw in enumerate(rows, start=2):
            try:
                vals = self._row_to_vals(raw, spec)
                prepared.append((row_num, vals))
            except UserError as err:
                errors.append({"row": row_num, "message": str(err)})

        if errors:
            return self._log_result(spec, filename, file_content, 0, 0, errors, failed=True)

        created = updated = 0
        try:
            with self.env.cr.savepoint():
                for row_num, vals in prepared:
                    existing = self._find_existing(model, spec, vals)
                    if existing:
                        existing.write(vals)
                        updated += 1
                    else:
                        model.create(vals)
                        created += 1
        except Exception as err:
            errors.append({"row": 0, "message": str(err)})
            return self._log_result(spec, filename, file_content, 0, 0, errors, failed=True)

        return self._log_result(spec, filename, file_content, created, updated, errors)

    @api.model
    def _resolve_target(self, target_key, require_create=True):
        spec = get_target(target_key)
        if not spec:
            raise UserError("نوع البيانات المحدد غير مدعوم.")
        if spec["model"] not in self.env.registry.models:
            raise UserError("نموذج البيانات غير متوفر في النظام.")
        if require_create:
            self.env[spec["model"]].check_access_rights("create")
        return spec

    @api.model
    def _read_rows(self, content, filename, spec):
        name = (filename or "").lower()
        if name.endswith(".csv"):
            return self._read_csv(content, spec)
        return self._read_xlsx(content, spec)

    @api.model
    def _read_csv(self, content, spec):
        text = content.decode("utf-8-sig") if isinstance(content, bytes) else content
        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)
        if not all_rows:
            return []
        header_map = self._map_headers(all_rows[0], spec)
        data = []
        for row in all_rows[1:]:
            if not any(cell not in (None, "") for cell in row):
                continue
            data.append(self._cells_to_dict(row, header_map, spec))
        return data

    @api.model
    def _read_xlsx(self, content, spec):
        from openpyxl import load_workbook  # noqa: PLC0415

        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = wb.active
        rows_iter = sheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return []
        header_map = self._map_headers([str(c or "") for c in header_row], spec)
        data = []
        for row in rows_iter:
            if not row or not any(c not in (None, "") for c in row):
                continue
            data.append(self._cells_to_dict(row, header_map, spec))
        return data

    @api.model
    def _map_headers(self, header_row, spec):
        normalized = {h.strip(): idx for idx, h in enumerate(header_row) if h and str(h).strip()}
        mapping = {}
        for col_spec in spec["columns"]:
            header = col_spec["header"]
            if header not in normalized:
                raise UserError("عمود مفقود في الملف: %s" % header)
            mapping[col_spec["field"]] = normalized[header]
        return mapping

    @api.model
    def _cells_to_dict(self, row, header_map, spec):
        raw = {}
        for col_spec in spec["columns"]:
            idx = header_map[col_spec["field"]]
            value = row[idx] if idx < len(row) else None
            if value is not None and str(value).strip() != "":
                raw[col_spec["field"]] = value
        return raw

    @api.model
    def _row_to_vals(self, raw, spec):
        vals = {}
        for col_spec in spec["columns"]:
            field_name = col_spec["field"]
            value = raw.get(field_name)
            if col_spec.get("required") and (value is None or str(value).strip() == ""):
                raise UserError("حقل مطلوب فارغ: %s" % col_spec["header"])

            if value is None or str(value).strip() == "":
                continue

            if col_spec.get("relation"):
                vals[field_name] = self._resolve_relation(col_spec, value)
            else:
                vals[field_name] = self._cast_value(self.env[spec["model"]], field_name, value)
        return vals

    @api.model
    def _resolve_relation(self, col_spec, value):
        comodel = col_spec["relation"]
        lookup_field = col_spec.get("relation_lookup") or "name"
        if comodel not in self.env.registry.models:
            raise UserError("نموذج مرتبط غير متوفر: %s" % comodel)
        lookup_val = str(value).strip()
        rec = self.env[comodel].sudo().search([(lookup_field, "=", lookup_val)], limit=1)
        if not rec and lookup_field != "name":
            rec = self.env[comodel].sudo().search([("name", "=", lookup_val)], limit=1)
        if not rec:
            raise UserError("لم يتم العثور على سجل مرتبط (%s=%s)" % (lookup_field, lookup_val))
        return rec.id

    @api.model
    def _cast_value(self, model, field_name, value):
        field = model._fields.get(field_name)
        if not field:
            return value
        if field.type == "integer":
            return int(float(value))
        if field.type == "float":
            return float(value)
        if field.type == "boolean":
            return str(value).strip().lower() in ("1", "true", "yes", "نعم")
        if field.type == "date" and hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")
        return str(value).strip() if value is not None else False

    @api.model
    def _find_existing(self, model, spec, vals):
        domain = []
        for mf in spec.get("match_fields") or []:
            if mf not in vals:
                return model.browse()
            domain.append((mf, "=", vals[mf]))
        if not domain:
            return model.browse()
        return model.search(domain, limit=1)

    @api.model
    def _log_result(self, spec, filename, content, created, updated, errors, failed=False):
        attachment = self.env["ir.attachment"].create({
            "name": filename or "import.xlsx",
            "datas": base64.b64encode(content if isinstance(content, bytes) else content.encode()),
            "res_model": "dashboard.import",
        })
        record = self.create({
            "name": "%s — %s" % (spec["label"], fields.Datetime.now()),
            "attachment_ids": [(6, 0, attachment.ids)],
            "target_key": spec["key"],
            "target_model": spec["model"],
            "target_label": spec["label"],
            "created_count": created,
            "updated_count": updated,
            "error_count": len(errors),
            "error_log": json.dumps(errors, ensure_ascii=False) if errors else "",
            "state": "failed" if failed else "processed",
            "note": "تم الاستيراد عبر معالج لوحة التحكم" if not failed else "فشل الاستيراد",
        })
        return {
            "import_id": record.id,
            "created": created,
            "updated": updated,
            "errors": errors,
            "error_count": len(errors),
            "target_label": spec["label"],
        }
